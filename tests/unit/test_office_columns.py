"""Office-marked --columns tests for xlsx/xlsm (piilint[office])."""

from __future__ import annotations

from pathlib import Path

import pytest
from typer.testing import CliRunner

pytest.importorskip("openpyxl")

from openpyxl import Workbook, load_workbook

from piilint.cli import app
from piilint.config import default_config
from piilint.engine import scan_path
from piilint.findings import EntityType
from piilint.redact import redact_tree

pytestmark = pytest.mark.office

COLUMNS_AGENT_NAMES = ["Alice Exampleton", "Bob Sampleton"]
COLUMNS_PHONES = ["2127350182", "4159032741"]
COLUMNS_SKILL_DIRTY = "416-735-0182"


def _write_columns_xlsx(path: Path) -> None:
    """Synthetic call-log: Agent names, numeric ANI/From, dirty Skill. Never real PII."""
    wb = Workbook()
    sheet = wb.active
    assert sheet is not None
    sheet.title = "Calls"
    sheet.append(["Agent", "ANI/From", "Skill"])
    sheet.append(["Alice Exampleton", 2127350182, COLUMNS_SKILL_DIRTY])
    sheet.append(["Bob Sampleton", 4159032741, "ops.beta@corpmail.test"])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def test_redact_xlsx_columns_headers(tmp_path: Path) -> None:
    src = tmp_path / "call_log.xlsx"
    _write_columns_xlsx(src)
    src_bytes = src.read_bytes()
    out = tmp_path / "clean"
    runner = CliRunner()
    result = runner.invoke(
        app,
        ["redact", str(src), "-o", str(out), "--columns", "Agent,ANI/From"],
    )
    assert result.exit_code == 0, result.output
    dest = out / "call_log.xlsx"
    assert dest.is_file()

    wb = load_workbook(dest, data_only=True)
    try:
        values: list[str] = []
        for row in wb.active.iter_rows():
            for cell in row:
                if cell.value is not None:
                    values.append(str(cell.value))
        blob = "\n".join(values)
        for raw in COLUMNS_PHONES:
            assert raw not in blob
        assert COLUMNS_SKILL_DIRTY in blob
        assert "ops.beta@corpmail.test" in blob
        assert wb.active["C2"].value == COLUMNS_SKILL_DIRTY
        for name in COLUMNS_AGENT_NAMES:
            assert name in blob
    finally:
        wb.close()
    assert src.read_bytes() == src_bytes


def test_redact_xlsx_columns_letter_same_as_header(tmp_path: Path) -> None:
    src = tmp_path / "call_log.xlsx"
    _write_columns_xlsx(src)
    out = tmp_path / "clean"
    result = redact_tree(src, out, config=default_config(), columns=["B"])
    assert result.files_written == 1
    dest = out / "call_log.xlsx"
    wb = load_workbook(dest, data_only=True)
    try:
        values = [str(c.value) for row in wb.active.iter_rows() for c in row if c.value is not None]
        blob = "\n".join(values)
        for raw in COLUMNS_PHONES:
            assert raw not in blob
        assert COLUMNS_SKILL_DIRTY in blob
        for name in COLUMNS_AGENT_NAMES:
            assert name in blob
        assert isinstance(wb.active["B2"].value, str)
    finally:
        wb.close()


def test_scan_xlsx_columns_agent_skips_ani_phones(tmp_path: Path) -> None:
    src = tmp_path / "call_log.xlsx"
    _write_columns_xlsx(src)
    full = scan_path(src)
    assert any(f.entity == EntityType.PHONE for f in full.findings)

    limited = scan_path(src, columns=["Agent"])
    assert not any(f.entity == EntityType.PHONE for f in limited.findings)

    runner = CliRunner()
    result = runner.invoke(app, ["scan", str(src), "--columns", "Agent", "--fail-on", "never"])
    assert result.exit_code == 0, result.output
    assert "PHONE" not in result.output


def test_scan_xlsx_columns_repeatable_and_sheet_scope(tmp_path: Path) -> None:
    src = tmp_path / "call_log.xlsx"
    _write_columns_xlsx(src)
    runner = CliRunner()
    result = runner.invoke(
        app,
        [
            "scan",
            str(src),
            "--columns",
            "Agent",
            "--columns",
            "Calls!B",
            "--fail-on",
            "never",
            "--format",
            "json",
        ],
    )
    assert result.exit_code == 0, result.output
    assert "PHONE" in result.output


def test_columns_unknown_exits_2_lists_headers(tmp_path: Path) -> None:
    src = tmp_path / "call_log.xlsx"
    _write_columns_xlsx(src)
    runner = CliRunner()
    result = runner.invoke(app, ["scan", str(src), "--columns", "NotAColumn"])
    assert result.exit_code == 2, result.output
    assert "Unknown column" in result.output
    assert "Agent" in result.output
    assert "ANI/From" in result.output
    assert "Skill" in result.output

    result = runner.invoke(
        app, ["redact", str(src), "-o", str(tmp_path / "out"), "--columns", "Nope"]
    )
    assert result.exit_code == 2, result.output
    assert "Unknown column" in result.output
    assert "Agent" in result.output


def test_columns_empty_sheet_exits_2(tmp_path: Path) -> None:
    src = tmp_path / "empty.xlsx"
    Workbook().save(src)
    runner = CliRunner()
    result = runner.invoke(app, ["scan", str(src), "--columns", "Agent"])
    assert result.exit_code == 2, result.output
    assert "Unknown column" in result.output or "No header" in result.output


def test_scan_xlsx_without_columns_still_finds_phones(tmp_path: Path) -> None:
    src = tmp_path / "call_log.xlsx"
    _write_columns_xlsx(src)
    result = scan_path(src)
    assert any(f.entity == EntityType.PHONE for f in result.findings)


def test_columns_directory_without_sheets_exits_2(tmp_path: Path) -> None:
    (tmp_path / "note.txt").write_text("hello 212-735-0182\n", encoding="utf-8")
    runner = CliRunner()
    result = runner.invoke(app, ["scan", str(tmp_path), "--columns", "Agent"])
    assert result.exit_code == 2, result.output
    assert "xlsx" in result.output.lower() or "spreadsheet" in result.output.lower()
