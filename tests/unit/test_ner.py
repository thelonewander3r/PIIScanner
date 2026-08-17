"""Unit tests for optional NER (PERSON/ADDRESS) — skip gracefully without extra/model."""

from __future__ import annotations

from pathlib import Path

import pytest
from typer.testing import CliRunner

from piilint.cli import app
from piilint.engine import scan_path
from piilint.findings import EntityType
from piilint.recognizers import build_default_registry
from piilint.recognizers.ner import (
    SPACY_MODEL,
    ner_deps_available,
    spacy_model_available,
)

CORPUS_NER = Path(__file__).resolve().parent.parent / "corpus" / "text" / "names_addresses.txt"

pytestmark = pytest.mark.ner


def test_ner_disabled_by_default() -> None:
    registry = build_default_registry()
    enabled = {r.entity for r in registry.enabled_recognizers()}
    assert EntityType.PERSON not in enabled
    assert EntityType.ADDRESS not in enabled

    result = scan_path(CORPUS_NER)
    entities = {f.entity for f in result.findings}
    assert EntityType.PERSON not in entities
    assert EntityType.ADDRESS not in entities


def test_cli_ner_missing_deps_exits_2(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr("piilint.recognizers.ner.ner_deps_available", lambda: False)
    runner = CliRunner()
    result = runner.invoke(app, ["scan", str(CORPUS_NER), "--ner", "--fail-on", "never"])
    assert result.exit_code == 2, result.output
    assert "piilint[ner]" in result.output
    assert "setup-ner" in result.output


def test_cli_ner_missing_model_exits_2(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr("piilint.recognizers.ner.ner_deps_available", lambda: True)
    monkeypatch.setattr(
        "piilint.recognizers.ner.spacy_model_available",
        lambda _m=SPACY_MODEL: False,
    )
    runner = CliRunner()
    result = runner.invoke(app, ["scan", str(CORPUS_NER), "--ner", "--fail-on", "never"])
    assert result.exit_code == 2, result.output
    assert "setup-ner" in result.output


def test_setup_ner_reports_missing_deps(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr("piilint.recognizers.ner.ner_deps_available", lambda: False)

    def _should_not_download(_model: str = SPACY_MODEL) -> None:
        raise AssertionError("download must not run when deps missing")

    monkeypatch.setattr("piilint.recognizers.ner.download_spacy_model", _should_not_download)
    runner = CliRunner()
    result = runner.invoke(app, ["setup-ner"])
    assert result.exit_code == 2, result.output
    assert "piilint[ner]" in result.output


@pytest.mark.skipif(not ner_deps_available(), reason="piilint[ner] not installed")
@pytest.mark.skipif(
    not spacy_model_available(),
    reason=f"{SPACY_MODEL} not installed; run piilint setup-ner",
)
def test_ner_finds_person_and_address() -> None:
    result = scan_path(CORPUS_NER, enable_ner=True)
    entities = {f.entity for f in result.findings}
    assert EntityType.PERSON in entities, result.findings
    assert EntityType.ADDRESS in entities, result.findings
    for finding in result.findings:
        if finding.entity in {EntityType.PERSON, EntityType.ADDRESS}:
            # Masked output only — full fake names must not appear in samples.
            assert "Alice Exampleton" not in finding.masked_sample
            assert "Fictitious Lane" not in finding.masked_sample


def test_cli_redact_ner_missing_deps_exits_2(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    monkeypatch.setattr("piilint.recognizers.ner.ner_deps_available", lambda: False)
    src = tmp_path / "note.txt"
    src.write_text("hello\n", encoding="utf-8")
    runner = CliRunner()
    result = runner.invoke(app, ["redact", str(src), "-o", str(tmp_path / "out"), "--ner"])
    assert result.exit_code == 2, result.output
    assert "piilint[ner]" in result.output


def test_cli_redact_ner_missing_model_exits_2(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    monkeypatch.setattr("piilint.recognizers.ner.ner_deps_available", lambda: True)
    monkeypatch.setattr(
        "piilint.recognizers.ner.spacy_model_available",
        lambda _m=SPACY_MODEL: False,
    )
    src = tmp_path / "note.txt"
    src.write_text("hello\n", encoding="utf-8")
    runner = CliRunner()
    result = runner.invoke(app, ["redact", str(src), "-o", str(tmp_path / "out"), "--ner"])
    assert result.exit_code == 2, result.output
    assert "setup-ner" in result.output


_XLSX_NAMES = ("Alice Exampleton", "Bob Sampleton")
_XLSX_PHONES = ("2127350182", "4159032741")


def _write_agent_xlsx(path: Path) -> None:
    """Synthetic call-log: Agent text + numeric phones. Never real PII."""
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    wb = Workbook()
    sheet = wb.active
    assert sheet is not None
    sheet.title = "Calls"
    sheet.append(["Agent", "Phone", "Duration"])
    sheet.append(["Alice Exampleton", 2127350182, 42])
    sheet.append(["Bob Sampleton", 4159032741, 17])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _xlsx_blob(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    try:
        values: list[str] = []
        for row in wb.active.iter_rows():
            for cell in row:
                if cell.value is not None:
                    values.append(str(cell.value))
        return "\n".join(values)
    finally:
        wb.close()


@pytest.mark.skipif(not ner_deps_available(), reason="piilint[ner] not installed")
@pytest.mark.skipif(
    not spacy_model_available(),
    reason=f"{SPACY_MODEL} not installed; run piilint setup-ner",
)
def test_redact_xlsx_ner_masks_person(tmp_path: Path) -> None:
    pytest.importorskip("openpyxl")
    from piilint.config import default_config
    from piilint.redact import redact_tree

    src = tmp_path / "agent_log.xlsx"
    _write_agent_xlsx(src)
    src_bytes = src.read_bytes()

    before = scan_path(src, enable_ner=True)
    person_hits = [f for f in before.findings if f.entity == EntityType.PERSON]
    assert person_hits, before.findings
    for finding in person_hits:
        for name in _XLSX_NAMES:
            assert name not in finding.masked_sample

    out = tmp_path / "clean"
    result = redact_tree(src, out, config=default_config(), enable_ner=True)
    assert result.files_written == 1
    assert result.spans_redacted > 0
    dest = out / "agent_log.xlsx"
    assert dest.is_file()

    blob = _xlsx_blob(dest)
    for name in _XLSX_NAMES:
        assert name not in blob
    for phone in _XLSX_PHONES:
        assert phone not in blob

    after = scan_path(dest, enable_ner=True)
    after_blob = "\n".join(
        str(part) for f in after.findings for part in (f.masked_sample, f.normalized_value) if part
    )
    for name in _XLSX_NAMES:
        assert name not in after_blob
    assert not any(f.entity == EntityType.PERSON for f in after.findings)
    assert not any(f.entity == EntityType.PHONE for f in after.findings)

    assert src.read_bytes() == src_bytes
    src_blob = _xlsx_blob(src)
    for name in _XLSX_NAMES:
        assert name in src_blob


@pytest.mark.skipif(not ner_deps_available(), reason="piilint[ner] not installed")
@pytest.mark.skipif(
    not spacy_model_available(),
    reason=f"{SPACY_MODEL} not installed; run piilint setup-ner",
)
def test_cli_redact_xlsx_ner(tmp_path: Path) -> None:
    pytest.importorskip("openpyxl")

    src = tmp_path / "agent_log.xlsx"
    _write_agent_xlsx(src)
    out = tmp_path / "out"
    runner = CliRunner()
    result = runner.invoke(app, ["redact", str(src), "-o", str(out), "--ner"])
    assert result.exit_code == 0, result.output
    dest = out / "agent_log.xlsx"
    assert dest.is_file()
    blob = _xlsx_blob(dest)
    for name in _XLSX_NAMES:
        assert name not in blob
        assert name not in result.output
    for phone in _XLSX_PHONES:
        assert phone not in blob
        assert phone not in result.output
