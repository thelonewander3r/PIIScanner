"""Excel column selectors for --columns (xlsx/xlsm only)."""

from __future__ import annotations

import re
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path
from typing import Any

SHEET_SUFFIXES = frozenset({".xlsx", ".xlsm"})
_LETTER_RE = re.compile(r"^[A-Za-z]+$")


class ColumnError(Exception):
    """Unknown or inapplicable --columns selector (CLI exit 2)."""


def parse_column_args(raw: list[str] | None) -> list[str]:
    """Split comma-separated / repeatable --columns values; strip; drop empties."""
    if not raw:
        return []
    out: list[str] = []
    seen: set[str] = set()
    for item in raw:
        for part in item.split(","):
            token = part.strip()
            if not token or token in seen:
                continue
            seen.add(token)
            out.append(token)
    return out


def is_sheet_path(path: Path) -> bool:
    return path.suffix.lower() in SHEET_SUFFIXES


def ensure_sheet_files(paths: Iterable[Path]) -> None:
    """Raise if --columns was set but the walk has no xlsx/xlsm files."""
    if not any(is_sheet_path(path) for path in paths):
        raise ColumnError(
            "--columns applies to Excel .xlsx/.xlsm only; no spreadsheet files found."
        )


def split_selector(selector: str) -> tuple[str | None, str]:
    """Return (optional sheet name, token) for ``Agent`` or ``Sheet1!H``."""
    if "!" not in selector:
        return None, selector.strip()
    sheet, _, token = selector.partition("!")
    sheet = sheet.strip()
    token = token.strip()
    return (sheet or None, token)


def _cell_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value != value:  # NaN
        return None
    text = str(value).strip()
    return text if text and text.lower() != "none" else None


@dataclass(frozen=True, slots=True)
class SheetCatalog:
    title: str
    headers: dict[int, str]
    max_col: int


def catalog_from_workbook(wb: Any) -> list[SheetCatalog]:
    """First-row headers + usable column count per sheet."""
    sheets: list[SheetCatalog] = []
    for sheet in wb.worksheets:
        headers: dict[int, str] = {}
        max_col = 0
        row1 = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if row1 is not None:
            for col_idx, raw in enumerate(row1, start=1):
                text = _cell_text(raw)
                if text:
                    headers[col_idx] = text
                    max_col = max(max_col, col_idx)
        dim = getattr(sheet, "max_column", None)
        max_row = getattr(sheet, "max_row", None)
        has_body = isinstance(max_row, int) and max_row > 1
        if isinstance(dim, int) and dim > 0 and (headers or has_body):
            max_col = max(max_col, dim)
        sheets.append(SheetCatalog(title=sheet.title, headers=headers, max_col=max_col))
    return sheets


def _unknown_message(unmatched: list[str], catalog: list[SheetCatalog]) -> str:
    from openpyxl.utils import get_column_letter

    headers: list[str] = []
    seen: set[str] = set()
    max_letter_col = 0
    sheet_names: list[str] = []
    for sheet in catalog:
        sheet_names.append(sheet.title)
        for col_idx in sorted(sheet.headers):
            header = sheet.headers[col_idx]
            if header not in seen:
                seen.add(header)
                headers.append(header)
        max_letter_col = max(max_letter_col, sheet.max_col)

    lines = [f"Unknown column(s): {', '.join(unmatched)}"]
    if headers:
        lines.append(f"Valid headers: {', '.join(headers)}")
    else:
        lines.append("No header row found.")
    if max_letter_col:
        letters = [get_column_letter(i) for i in range(1, max_letter_col + 1)]
        lines.append(f"Column letters: {', '.join(letters)}")
    if sheet_names:
        lines.append("Sheets: " + ", ".join(sheet_names) + " (sheet-scoped form: SheetName!H)")
    return "\n".join(lines)


def resolve_workbook_columns(wb: Any, selectors: list[str]) -> set[tuple[str, int]]:
    """Map selectors to ``(sheet_title, 1-based col)``. Raises ColumnError if any miss."""
    from openpyxl.utils import get_column_letter

    catalog = catalog_from_workbook(wb)
    selected: set[tuple[str, int]] = set()
    unmatched: list[str] = []

    for selector in selectors:
        sheet_name, token = split_selector(selector)
        if not token:
            unmatched.append(selector)
            continue
        sheets = catalog
        if sheet_name is not None:
            sheets = [s for s in catalog if s.title == sheet_name]
            if not sheets:
                lowered = sheet_name.lower()
                sheets = [s for s in catalog if s.title.lower() == lowered]
            if not sheets:
                unmatched.append(selector)
                continue

        matched_here: set[tuple[str, int]] = set()
        want_letter = token.upper() if _LETTER_RE.fullmatch(token) else None
        for sheet in sheets:
            for col_idx, header in sheet.headers.items():
                if header == token:
                    matched_here.add((sheet.title, col_idx))
            if want_letter and sheet.max_col:
                for col_idx in range(1, sheet.max_col + 1):
                    if get_column_letter(col_idx) == want_letter:
                        matched_here.add((sheet.title, col_idx))
        if not matched_here:
            unmatched.append(selector)
        else:
            selected.update(matched_here)

    if unmatched:
        raise ColumnError(_unknown_message(unmatched, catalog))
    return selected
