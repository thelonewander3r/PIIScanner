"""Excel (.xlsx / .xlsm) adapter — optional ``piilint[office]`` (openpyxl)."""

from __future__ import annotations

import importlib.util
import sys
from collections.abc import Iterator
from pathlib import Path

from piilint.adapters import Unit

_INSTALL_HINT = 'pip install "piilint[office]"'
_warned = False


def office_xlsx_available() -> bool:
    return importlib.util.find_spec("openpyxl") is not None


def _warn_missing_once() -> None:
    global _warned
    if _warned:
        return
    _warned = True
    print(
        f"piilint: skipping .xlsx/.xlsm (openpyxl not installed). Install with: {_INSTALL_HINT}",
        file=sys.stderr,
    )


def _cell_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value != value:  # NaN
        return None
    text = str(value).strip()
    return text if text and text.lower() != "none" else None


class XlsxAdapter:
    name = "xlsx"
    extensions = frozenset({".xlsx", ".xlsm"})

    def supports(self, path: Path) -> bool:
        return path.suffix.lower() in self.extensions

    def iter_units(
        self,
        path: Path,
        *,
        rel_path: str,
        columns: list[str] | None = None,
    ) -> Iterator[Unit]:
        if not office_xlsx_available():
            if columns:
                from piilint.columns import ColumnError

                raise ColumnError(
                    "--columns requires piilint[office]. "
                    'Install with: pip install "piilint[office]"'
                )
            _warn_missing_once()
            return
        from openpyxl import load_workbook

        # read_only cannot rewind after cataloging first rows for --columns.
        try:
            wb = load_workbook(path, read_only=not bool(columns), data_only=True)
        except OSError:
            return
        except Exception:  # noqa: BLE001 — corrupt workbook
            return

        try:
            selected: set[tuple[str, int]] | None = None
            if columns:
                from piilint.columns import resolve_workbook_columns

                selected = resolve_workbook_columns(wb, columns)
            for sheet in wb.worksheets:
                headers: dict[int, str] = {}
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    for col_idx, raw in enumerate(row, start=1):
                        if selected is not None and (sheet.title, col_idx) not in selected:
                            continue
                        text = _cell_text(raw)
                        if text is None:
                            continue
                        if row_idx == 1:
                            headers[col_idx] = text
                        header = headers.get(col_idx)
                        context = header or sheet.title
                        yield Unit(
                            text=text,
                            path=rel_path,
                            column=f"{sheet.title}!{header}" if header else sheet.title,
                            row=row_idx,
                            context_key=context,
                            aggregate=False,
                        )
        finally:
            wb.close()
